"""
REMIC Data Loader - Parse Ginnie Mae REMIC monthly issuance data files.
"""
import os
import zipfile
import tempfile
from dataclasses import dataclass, field
from typing import Optional
from pathlib import Path

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    import pandas as pd
except ImportError:
    pd = None


@dataclass
class REMICDeal:
    """Parsed REMIC deal from issuance data."""
    series: str
    dealer: str
    trustee: str
    deal_type: str              # Single Family, Multifamily, Reverse REMIC
    groups: list[dict] = field(default_factory=list)
    total_bond_issuance: float = 0.0
    total_notional: float = 0.0
    month: str = ""
    year: int = 0

    @property
    def structure_summary(self) -> str:
        """Summarize the principal/interest structure types across groups."""
        prin_types = set()
        int_types = set()
        for g in self.groups:
            for t in g.get('principal_type', '').split('/'):
                if t.strip():
                    prin_types.add(t.strip())
            for t in g.get('interest_type', '').split('/'):
                if t.strip():
                    int_types.add(t.strip())
        return f"Principal: {'/'.join(sorted(prin_types))} | Interest: {'/'.join(sorted(int_types))}"

    def to_dict(self) -> dict:
        return {
            'series': self.series,
            'dealer': self.dealer,
            'trustee': self.trustee,
            'deal_type': self.deal_type,
            'total_bond_issuance': self.total_bond_issuance,
            'total_notional': self.total_notional,
            'groups': self.groups,
            'structure_summary': self.structure_summary,
            'period': f"{self.month} {self.year}",
        }


def parse_remic_xlsx(filepath: str) -> list[REMICDeal]:
    """Parse a REMIC monthly issuance xlsx file."""
    if openpyxl is None:
        raise ImportError("openpyxl required: pip install openpyxl")

    wb = openpyxl.load_workbook(filepath, data_only=True)
    deals = []

    # Find the issuance summary sheet
    target_sheet = None
    for name in wb.sheetnames:
        if 'Issuance Summary' in name:
            target_sheet = wb[name]
            break

    if target_sheet is None:
        target_sheet = wb.active

    # Extract month/year from title
    month_str = ""
    year_val = 0
    for row in target_sheet.iter_rows(min_row=1, max_row=3, values_only=True):
        for cell in row:
            if cell and 'REMIC Issuance' in str(cell):
                parts = str(cell).split()
                if len(parts) >= 2:
                    month_str = parts[0]
                    for p in parts:
                        try:
                            y = int(p)
                            if 2019 <= y <= 2030:
                                year_val = y
                        except ValueError:
                            continue
                break

    # Find header row
    header_row = None
    for row_idx, row in enumerate(target_sheet.iter_rows(min_row=1, max_row=20, values_only=True), 1):
        cells = [str(c).strip() if c else '' for c in row]
        if 'Series' in cells and 'Dealer' in cells:
            header_row = row_idx
            break

    if header_row is None:
        return deals

    # Parse deal rows
    current_deal = None
    for row in target_sheet.iter_rows(min_row=header_row + 1, max_row=500, values_only=True):
        cells = list(row)
        if len(cells) < 13:
            continue

        # cells[1] = Series, cells[2] = Dealer, cells[3] = Trustee,
        # cells[4] = Deal Type, cells[5] = Group Number,
        # cells[6] = Collateral Type, cells[7] = Coupon,
        # cells[8] = Original Term, cells[9] = Structure Principal Type,
        # cells[10] = Structure Interest Type, cells[11] = Bond Issuance Amount,
        # cells[12] = Bond Notional Amount
        series = str(cells[1]).strip() if cells[1] else ''
        dealer = str(cells[2]).strip() if cells[2] else ''

        # Skip total rows and empty rows
        if 'Total' in series or 'Grand' in series:
            if current_deal is not None:
                # Calculate totals
                current_deal.total_bond_issuance = sum(
                    g.get('bond_issuance', 0) for g in current_deal.groups)
                current_deal.total_notional = sum(
                    g.get('notional', 0) for g in current_deal.groups)
                deals.append(current_deal)
                current_deal = None
            continue

        if not series and not cells[5]:  # No series and no group number
            continue

        if series and not series.startswith('='):
            # New deal
            if current_deal is not None:
                current_deal.total_bond_issuance = sum(
                    g.get('bond_issuance', 0) for g in current_deal.groups)
                current_deal.total_notional = sum(
                    g.get('notional', 0) for g in current_deal.groups)
                deals.append(current_deal)

            current_deal = REMICDeal(
                series=series,
                dealer=dealer if dealer and dealer != '\xa0' else '',
                trustee=str(cells[3]).strip() if cells[3] and str(cells[3]).strip() != '\xa0' else '',
                deal_type=str(cells[4]).strip() if cells[4] and str(cells[4]).strip() != '\xa0' else '',
                month=month_str,
                year=year_val,
            )

        if current_deal is not None and cells[5]:
            group_num = cells[5]
            if str(group_num).strip() in ('', '\xa0'):
                continue
            group = {
                'group_number': str(group_num).strip(),
                'collateral_type': str(cells[6]).strip() if cells[6] else '',
                'coupon': float(cells[7]) if cells[7] and str(cells[7]).replace('.', '').isdigit() else 0.0,
                'original_term': int(float(str(cells[8]))) if cells[8] and str(cells[8]).replace('.', '').isdigit() else 0,
                'principal_type': str(cells[9]).strip() if cells[9] else '',
                'interest_type': str(cells[10]).strip() if cells[10] else '',
                'bond_issuance': float(cells[11]) if cells[11] and str(cells[11]).replace('.', '').isdigit() else 0.0,
                'notional': float(cells[12]) if cells[12] and str(cells[12]).replace('.', '').isdigit() else 0.0,
            }
            try:
                if cells[7]:
                    group['coupon'] = float(cells[7])
            except (ValueError, TypeError):
                pass
            try:
                if cells[11]:
                    group['bond_issuance'] = float(cells[11])
            except (ValueError, TypeError):
                pass
            try:
                if cells[12]:
                    group['notional'] = float(cells[12])
            except (ValueError, TypeError):
                pass

            current_deal.groups.append(group)

    # Don't forget last deal
    if current_deal is not None and current_deal.groups:
        current_deal.total_bond_issuance = sum(
            g.get('bond_issuance', 0) for g in current_deal.groups)
        current_deal.total_notional = sum(
            g.get('notional', 0) for g in current_deal.groups)
        deals.append(current_deal)

    return deals


def parse_remic_zip(filepath: str) -> list[REMICDeal]:
    """Parse a REMIC zip file containing xlsx data."""
    deals = []
    with tempfile.TemporaryDirectory() as tmpdir:
        with zipfile.ZipFile(filepath, 'r') as z:
            z.extractall(tmpdir)
            for f in os.listdir(tmpdir):
                if f.endswith('.xlsx') or f.endswith('.xls'):
                    try:
                        deals.extend(parse_remic_xlsx(os.path.join(tmpdir, f)))
                    except Exception:
                        continue
    return deals


def load_all_remic_data(data_dir: str) -> list[REMICDeal]:
    """Load all REMIC issuance data from a directory."""
    all_deals = []
    data_path = Path(data_dir)

    for f in sorted(data_path.iterdir()):
        if f.suffix == '.xlsx':
            try:
                deals = parse_remic_xlsx(str(f))
                all_deals.extend(deals)
            except Exception as e:
                print(f"Warning: Could not parse {f.name}: {e}")
        elif f.suffix == '.zip':
            try:
                deals = parse_remic_zip(str(f))
                all_deals.extend(deals)
            except Exception as e:
                print(f"Warning: Could not parse {f.name}: {e}")

    return all_deals


def analyze_remic_trends(deals: list[REMICDeal]) -> dict:
    """Analyze trends in REMIC issuance data."""
    if not deals:
        return {}

    total_issuance = sum(d.total_bond_issuance for d in deals)
    by_dealer = {}
    by_type = {}
    structure_types = {}
    coupons = []

    for d in deals:
        # By dealer
        if d.dealer:
            by_dealer[d.dealer] = by_dealer.get(d.dealer, 0) + d.total_bond_issuance

        # By deal type
        if d.deal_type:
            by_type[d.deal_type] = by_type.get(d.deal_type, 0) + d.total_bond_issuance

        # Structure types
        for g in d.groups:
            pt = g.get('principal_type', '')
            if pt:
                for t in pt.split('/'):
                    t = t.strip()
                    if t:
                        structure_types[t] = structure_types.get(t, 0) + 1

            if g.get('coupon', 0) > 0:
                coupons.append(g['coupon'])

    top_dealers = sorted(by_dealer.items(), key=lambda x: -x[1])[:10]

    return {
        'total_deals': len(deals),
        'total_issuance': total_issuance,
        'by_deal_type': by_type,
        'top_dealers': dict(top_dealers),
        'structure_type_frequency': structure_types,
        'coupon_range': (min(coupons), max(coupons)) if coupons else (0, 0),
        'avg_coupon': sum(coupons) / len(coupons) if coupons else 0,
    }
